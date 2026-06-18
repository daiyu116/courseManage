# SPDX-License-Identifier: AGPL-3.0-only
# Copyright (C) 2024-2026 CourseArrange Contributors
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
from database import get_db
from models import StudentFee, FeeLog, Student, Course, Schedule, Teacher, Class, Room, schedule_student, Settings
from schemas import (
    StudentFeeCreate, StudentFeeUpdate, StudentFee as StudentFeeSchema,
    FeeLogCreate, FeeLog as FeeLogSchema,
    PaymentRequest, RefundRequest, AlertThresholdUpdate, PaginatedFeeResponse
)
from routers.auth import get_current_user,get_current_system_admin_user, User
from datetime import datetime, date
from openpyxl import Workbook
from io import BytesIO
from fastapi.responses import StreamingResponse
from utils.logger import log_operation
from utils.wechat_notifier import wechat_notifier
from sqlalchemy import select
import json

router = APIRouter()

def get_hours_per_lesson(db: Session) -> float:
    settings = db.query(Settings).first()
    if settings and settings.hours_per_lesson is not None:
        return settings.hours_per_lesson
    return 2.0

def check_fee_manager_permission(db: Session, current_user: User):
    """检查用户是否有费用管理权限（管理员或费用管理导师）+ License 授权"""
    from routers.license import _check_premium_feature
    if not _check_premium_feature('fee_management', db):
        raise HTTPException(status_code=403, detail="费用管理功能需要购买授权后才能使用")
    # 1. 如果是超级管理员或系统管理员，直接通过
    if current_user.role in ['super_admin', 'system_admin']:
        return True
    
    # 2. 检查是否是费用管理导师
    if current_user.teacher_id:
        settings = db.query(Settings).first()
        if settings and settings.fee_managers:
            try:
                fee_managers = json.loads(settings.fee_managers)
                if current_user.teacher_id in fee_managers:
                    return True
            except:
                pass
    return False

@router.get("/student-fees", response_model=PaginatedFeeResponse)
def get_student_fees(
    skip: int = 0,
    limit: int = 100,
    student_id: Optional[int] = None,
    course_id: Optional[int] = None,
    is_active: Optional[bool] = None,
    search: Optional[str] = None,
    sort_field: Optional[str] = None,
    sort_order: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")
    
    """获取学生课时费记录列表"""
    query = db.query(StudentFee).options(
        joinedload(StudentFee.student).joinedload(Student.classes),
        joinedload(StudentFee.course)
    )

    # 搜索逻辑：如果提供了search参数，则搜索学员姓名或科目名称包含该关键词的记录
    if search:
        # 搜索学员姓名或科目名称
        query = query.join(Student).join(Course).filter(
            (Student.name.contains(search)) | (Course.name.contains(search))
        )

    if student_id is not None:
        query = query.filter(StudentFee.student_id == student_id)
    if course_id is not None:
        query = query.filter(StudentFee.course_id == course_id)
    if is_active is not None:
        query = query.filter(StudentFee.is_active == is_active)
    
    # 应用排序
    if sort_field:
        sort_column = getattr(StudentFee, sort_field, None)
        if sort_column:
            if sort_order == "desc":
                query = query.order_by(sort_column.desc())
            else:
                query = query.order_by(sort_column.asc())
    
    total = query.count()
    fees = query.offset(skip).limit(limit).all()
    
    result = []
    for fee in fees:
        student = fee.student
        course = fee.course
        
        # 获取学员的班级信息
        student_classes = []
        if student and student.classes:
            for class_ in student.classes:
                student_classes.append({
                    "id": class_.id,
                    "name": class_.name
                })
        
        # 获取科目的导师信息
        course_teachers = []
        if course and course.teachers:
            for teacher in course.teachers:
                course_teachers.append({
                    "id": teacher.id,
                    "name": teacher.name,
                    "contact_phone": teacher.contact_phone,
                    "email": teacher.email
                })
        # 重新计算总实收金额：所有收费记录的实缴金额总和
        payment_logs = db.query(FeeLog).filter(
            FeeLog.student_id == fee.student_id,
            FeeLog.course_id == fee.course_id,
            FeeLog.log_type == "payment"
        ).all()
        calculated_total_actual_amount = sum(log.amount for log in payment_logs)
    
        # 计算总应收金额：所有收费记录的应收金额总和
        total_receivable_amount = sum(log.receivable_amount for log in payment_logs)
        
        # 计算总退费金额：所有退费记录的金额总和
        refund_logs = db.query(FeeLog).filter(
            FeeLog.student_id == fee.student_id,
            FeeLog.course_id == fee.course_id,
            FeeLog.log_type == "refund"
        ).all()
        total_refund_amount = sum(abs(log.amount) for log in refund_logs)
        
        # 计算已消耗课时
        consumed_logs = db.query(FeeLog).filter(
            FeeLog.student_id == fee.student_id,
            FeeLog.course_id == fee.course_id,
            FeeLog.log_type == "consume"
        ).all()
        consumed_hours = sum(log.hours for log in consumed_logs)
        
        # 获取消耗的课程安排列表（用于tooltip显示）
        consumed_schedules = []
        for log in consumed_logs:
            if log.schedule_id:
                schedule = db.query(Schedule).filter(Schedule.id == log.schedule_id).first()
                if schedule:
                    teacher = db.query(Teacher).filter(Teacher.id == schedule.teacher_id).first()
                    class_ = db.query(Class).filter(Class.id == schedule.class_id).first()
                    room = db.query(Room).filter(Room.id == schedule.room_id).first()
                    
                    consumed_schedules.append({
                        "date": schedule.start_date.strftime("%Y-%m-%d") if schedule.start_date else "",
                        "time": f"{schedule.start_time}-{schedule.end_time}",
                        "teacher": teacher.name if teacher else "",
                        "class": class_.name if class_ else "",
                        "room": room.name if room else "",
                        "hours": log.hours
                    })
        
        result.append(StudentFeeSchema(
            id=fee.id,
            student_id=fee.student_id,
            course_id=fee.course_id,
            start_date=fee.start_date,
            hourly_fee=fee.hourly_fee,
            total_receivable_amount=total_receivable_amount,
            total_actual_amount=calculated_total_actual_amount,
            total_refund_amount=total_refund_amount,
            remaining_hours=fee.remaining_hours,
            consumed_hours=consumed_hours,
            consumed_schedules=consumed_schedules,
            alert_threshold=fee.alert_threshold,
            is_active=fee.is_active,
            student_name=student.name if student else "",
            student_school=student.school if student else "",
            student_grade=student.grade if student else "",
            student_contact_person=student.contact_person if student else "",
            student_contact_phone=student.contact_phone if student else "",
            student_classes=student_classes,
            student_is_active=student.is_active if student else True,
            course_name=course.name if course else "",
            course_teachers=course_teachers,
            created_at=fee.created_at,
            updated_at=fee.updated_at
        ))
    
    return {
        "items": result,
        "total": total
    }

@router.get("/student-fees/{fee_id}", response_model=StudentFeeSchema)
def get_student_fee(
    fee_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")

    """获取单个学生课时费记录"""
    fee = db.query(StudentFee).filter(StudentFee.id == fee_id).first()
    if not fee:
        log_operation(db, "费用管理", "获取课时费详情失败", f"课时费记录ID {fee_id} 不存在", current_user.username, "WARNING")
        raise HTTPException(status_code=404, detail="课时费记录不存在")
    
    student = db.query(Student).filter(Student.id == fee.student_id).first()
    course = db.query(Course).filter(Course.id == fee.course_id).first()
    
    # 重新计算总实收金额：所有收费记录的实缴金额总和
    payment_logs = db.query(FeeLog).filter(
        FeeLog.student_id == fee.student_id,
        FeeLog.course_id == fee.course_id,
        FeeLog.log_type == "payment"
    ).all()
    calculated_total_actual_amount = sum(log.amount for log in payment_logs)
    
    # 计算总应收金额：所有收费记录的应收金额总和
    total_receivable_amount = sum(log.receivable_amount for log in payment_logs)
    
    # 计算总退费金额：所有退费记录的金额总和
    refund_logs = db.query(FeeLog).filter(
        FeeLog.student_id == fee.student_id,
        FeeLog.course_id == fee.course_id,
        FeeLog.log_type == "refund"
    ).all()
    total_refund_amount = sum(abs(log.amount) for log in refund_logs)
    
    # 计算已消耗课时
    consumed_logs = db.query(FeeLog).filter(
        FeeLog.student_id == fee.student_id,
        FeeLog.course_id == fee.course_id,
        FeeLog.log_type == "consume"
    ).all()
    consumed_hours = sum(log.hours for log in consumed_logs)
    
    # 获取消耗的课程安排列表
    consumed_schedules = []
    for log in consumed_logs:
        if log.schedule_id:
            schedule = db.query(Schedule).filter(Schedule.id == log.schedule_id).first()
            if schedule:
                teacher = db.query(Teacher).filter(Teacher.id == schedule.teacher_id).first()
                class_ = db.query(Class).filter(Class.id == schedule.class_id).first()
                room = db.query(Room).filter(Room.id == schedule.room_id).first()
                
                consumed_schedules.append({
                    "date": schedule.start_date.strftime("%Y-%m-%d") if schedule.start_date else "",
                    "time": f"{schedule.start_time}-{schedule.end_time}",
                    "teacher": teacher.name if teacher else "",
                    "class": class_.name if class_ else "",
                    "room": room.name if room else "",
                    "hours": log.hours
                })
    
    return StudentFeeSchema(
        id=fee.id,
        student_id=fee.student_id,
        course_id=fee.course_id,
        start_date=fee.start_date,
        hourly_fee=fee.hourly_fee,
        total_receivable_amount=total_receivable_amount,
        total_actual_amount=calculated_total_actual_amount,
        total_refund_amount=total_refund_amount,
        remaining_hours=fee.remaining_hours,
        consumed_hours=consumed_hours,
        consumed_schedules=consumed_schedules,
        alert_threshold=fee.alert_threshold,
        is_active=fee.is_active,
        student_name=student.name if student else "",
        course_name=course.name if course else "",
        created_at=fee.created_at,
        updated_at=fee.updated_at
    )

@router.post("/student-fees", response_model=StudentFeeSchema)
def create_student_fee(
    fee: StudentFeeCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")

    """创建学生课时费记录"""    
    student = db.query(Student).filter(Student.id == fee.student_id).first()
    if not student:
        log_operation(db, "费用管理", "创建课时费失败", f"学生ID {fee.student_id} 不存在", current_user.username, "WARNING")
        raise HTTPException(status_code=404, detail="学生不存在")
    
    course = db.query(Course).filter(Course.id == fee.course_id).first()
    if not course:
        log_operation(db, "费用管理", "创建课时费失败", f"科目ID {fee.course_id} 不存在", current_user.username, "WARNING")
        raise HTTPException(status_code=404, detail="科目不存在")
    
    existing_fee = db.query(StudentFee).filter(
        StudentFee.student_id == fee.student_id,
        StudentFee.course_id == fee.course_id
    ).first()
    if existing_fee:
        log_operation(db, "费用管理", "创建课时费失败", f"学生 {student.name} - 科目 {course.name} 已有课时费记录", current_user.username, "WARNING")
        raise HTTPException(status_code=400, detail="该学生该科目已存在课时费记录")
    
    # 计算本次应收金额 = 课节数 * 小时数每课节 * (课时费/小时)
    hpl = get_hours_per_lesson(db)
    receivable_amount = fee.lesson_count * hpl * fee.hourly_fee
    actual_amount = receivable_amount - fee.discount_amount
    remaining_hours = fee.lesson_count * hpl
    
    db_fee = StudentFee(
        student_id=fee.student_id,
        course_id=fee.course_id,
        start_date=fee.start_date,
        hourly_fee=fee.hourly_fee,
        total_receivable_amount=receivable_amount,
        total_actual_amount=actual_amount,
        total_refund_amount=0.0,
        total_lesson_count=fee.lesson_count,
        consumed_hours=0.0,
        remaining_hours=remaining_hours,
        alert_threshold=fee.alert_threshold,
        is_active=fee.is_active
    )
    db.add(db_fee)
    db.commit()
    
    # 创建收费日志
    payment_method_desc = f"，{fee.payment_method}" if fee.payment_method else ""
    log = FeeLog(
        student_id=fee.student_id,
        course_id=fee.course_id,
        log_type="payment",
        amount=actual_amount,
        receivable_amount=receivable_amount,
        hours=0,
        remaining_amount=actual_amount,
        remaining_hours=remaining_hours,
        payment_date=fee.payment_date or datetime.now(),
        description=f"初始收费（{fee.lesson_count}节课，{fee.lesson_count * hpl}小时）{'，优惠' + str(fee.discount_amount) + '元' if fee.discount_amount > 0 else ''}{'，收费来自' +payment_method_desc}"
    )
    db.add(log)
    db.commit()
    
    # 检查是否有之前完训的课程安排，自动计算消耗
    completed_schedules = db.query(Schedule).filter(
        Schedule.course_id == fee.course_id,
        Schedule.execution_status == "completed"
    ).all()
    
    has_consumed = False  # 标记是否有消耗记录被创建
    
    for schedule in completed_schedules:
        # 只处理起算日期及之后的完训课程（包括当天）
        # schedule.start_date 是 Date 类型，fee.start_date 是 DateTime 类型
        if schedule.start_date and fee.start_date:
            schedule_date = schedule.start_date  # 已经是 date 类型
            fee_date = fee.start_date.date() if hasattr(fee.start_date, 'date') else fee.start_date  # 从 datetime 提取 date
            
            # 如果课程安排在起算日期之前，跳过
            if schedule_date < fee_date:
                continue
        
        # 直接查询该课程安排的参与学员列表（通过schedule_student表）
        stmt = select(schedule_student.c.student_id, schedule_student.c.attendance_status).where(
            schedule_student.c.schedule_id == schedule.id
        )
        result = db.execute(stmt)
        scheduled_student_info = {row[0]: row[1] for row in result.fetchall()}
        if fee.student_id in scheduled_student_info and scheduled_student_info[fee.student_id] == 'present':
            # 检查是否已经创建了消耗日志
            existing_log = db.query(FeeLog).filter(
                FeeLog.student_id == fee.student_id,
                FeeLog.course_id == fee.course_id,
                FeeLog.schedule_id == schedule.id,
                FeeLog.log_type == "consume"
            ).first()
            
            if not existing_log:
                # 创建消耗日志
                start_time = datetime.strptime(schedule.start_time, "%H:%M")
                end_time = datetime.strptime(schedule.end_time, "%H:%M")
                hours = (end_time - start_time).total_seconds() / 3600
                
                db_fee.consumed_hours += hours
                db_fee.remaining_hours = db_fee.total_lesson_count * hpl - db_fee.consumed_hours
                
                consume_log = FeeLog(
                    student_id=fee.student_id,
                    course_id=schedule.course_id,
                    schedule_id=schedule.id,
                    log_type="consume",
                    amount=-hours * db_fee.hourly_fee,
                    hours=hours,
                    remaining_amount=db_fee.total_actual_amount - db_fee.total_refund_amount,
                    remaining_hours=db_fee.remaining_hours,
                    description=f"完训消耗 {hours} 小时(根据规则和实际情况触发记录)"
                )
                db.add(consume_log)
                has_consumed = True
    
    # 如果有消耗记录被创建，需要提交db_fee的更新
    if has_consumed:
        db.commit()
        db.refresh(db_fee)

    log_operation(db, "费用管理", "新增", f"成功新增课费项：{student.name} - {course.name}",
        user=current_user.username
    )
    db.refresh(db_fee)
    
    return StudentFeeSchema(
        id=db_fee.id,
        student_id=db_fee.student_id,
        course_id=db_fee.course_id,
        start_date=db_fee.start_date,
        hourly_fee=db_fee.hourly_fee,
        total_receivable_amount=db_fee.total_receivable_amount,
        total_actual_amount=db_fee.total_actual_amount,
        total_refund_amount=db_fee.total_refund_amount,
        total_lesson_count=db_fee.total_lesson_count,
        consumed_hours=db_fee.consumed_hours,
        remaining_hours=db_fee.remaining_hours,
        alert_threshold=db_fee.alert_threshold,
        is_active=db_fee.is_active,
        student_name=student.name,
        course_name=course.name,
        created_at=db_fee.created_at,
        updated_at=db_fee.updated_at
    )

@router.get("/debug-auto-consume/{student_id}/{course_id}")
def debug_auto_consume(
    student_id: int,
    course_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """调试自动消耗逻辑 - 返回详细的检查信息"""
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足")
    
    result = {
        "student_id": student_id,
        "course_id": course_id,
        "checks": []
    }
    
    # 1. 查询该科目下所有完训的课程安排
    completed_schedules = db.query(Schedule).filter(
        Schedule.course_id == course_id,
        Schedule.execution_status == "completed"
    ).all()
    
    result["checks"].append({
        "step": "1. 查询完训课程安排",
        "count": len(completed_schedules),
        "schedules": [
            {
                "id": s.id,
                "start_date": str(s.start_date),
                "execution_status": s.execution_status,
                "class_id": s.class_id
            } for s in completed_schedules
        ]
    })
    
    # 2. 查询学员的课费项
    fee = db.query(StudentFee).filter(
        StudentFee.student_id == student_id,
        StudentFee.course_id == course_id
    ).first()
    
    if not fee:
        result["checks"].append({
            "step": "2. 查询课费项",
            "found": False,
            "message": f"未找到学员ID {student_id} - 科目ID {course_id} 的课费项"
        })
        return result
    
    result["checks"].append({
        "step": "2. 查询课费项",
        "found": True,
        "fee_id": fee.id,
        "start_date": str(fee.start_date),
        "consumed_hours": fee.consumed_hours,
        "remaining_hours": fee.remaining_hours,
        "total_lesson_count": fee.total_lesson_count
    })
    
    # 3. 检查每个完训课程安排
    for schedule in completed_schedules:
        check_result = {
            "schedule_id": schedule.id,
            "start_date": str(schedule.start_date),
            "fee_start_date": str(fee.start_date),
            "date_check": None,
            "student_in_schedule": None,
            "existing_log": None,
            "should_consume": False
        }
        
        # 日期检查 - 统一转换为date类型再比较
        # schedule.start_date 是 Date 类型，fee.start_date 是 DateTime 类型
        if schedule.start_date and fee.start_date:
            schedule_date = schedule.start_date  # 已经是 date 类型
            fee_date = fee.start_date.date() if hasattr(fee.start_date, 'date') else fee.start_date  # 从 datetime 提取 date
            
            if schedule_date < fee_date:
                check_result["date_check"] = "跳过（课程安排在起算日期之前）"
            else:
                check_result["date_check"] = "通过（课程安排在起算日期及之后）"
        else:
            check_result["date_check"] = "跳过（日期为空）"
        
        # 检查学员是否在课程安排中
        stmt = select(schedule_student.c.student_id, schedule_student.c.attendance_status).where(
            schedule_student.c.schedule_id == schedule.id
        )
        query_result = db.execute(stmt)
        scheduled_student_info = {row[0]: row[1] for row in query_result.fetchall()}
        
        check_result["student_in_schedule"] = {
            "in_list": student_id in scheduled_student_info,
            "attendance_status": scheduled_student_info.get(student_id),
            "scheduled_students": list(scheduled_student_info.keys())
        }
        
        # 检查是否已有消耗日志
        existing_log = db.query(FeeLog).filter(
            FeeLog.student_id == student_id,
            FeeLog.course_id == course_id,
            FeeLog.schedule_id == schedule.id,
            FeeLog.log_type == "consume"
        ).first()
        
        check_result["existing_log"] = {
            "exists": existing_log is not None,
            "log_id": existing_log.id if existing_log else None
        }
        
        # 判断是否应该消耗
        if (check_result["date_check"] and "通过" in check_result["date_check"] and
            check_result["student_in_schedule"]["in_list"] and
            check_result["student_in_schedule"]["attendance_status"] == 'present' and
            not check_result["existing_log"]["exists"]):
            check_result["should_consume"] = True
        
        result["checks"].append({
            "step": f"3. 检查课程安排ID {schedule.id}",
            **check_result
        })
    
    return result

@router.post("/trigger-auto-consume/{fee_id}")
def trigger_auto_consume(
    fee_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """手动触发自动消耗检查"""
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足")
    
    fee = db.query(StudentFee).filter(StudentFee.id == fee_id).first()
    if not fee:
        raise HTTPException(status_code=404, detail="课费项不存在")
    
    student = db.query(Student).filter(Student.id == fee.student_id).first()
    course = db.query(Course).filter(Course.id == fee.course_id).first()
    
    # 查询该科目下所有完训的课程安排
    completed_schedules = db.query(Schedule).filter(
        Schedule.course_id == fee.course_id,
        Schedule.execution_status == "completed"
    ).all()
    
    consumed_count = 0
    
    hpl_consume = get_hours_per_lesson(db)
    for schedule in completed_schedules:
        # 只处理起算日期及之后的完训课程（包括当天）
        if schedule.start_date and fee.start_date:
            schedule_date = schedule.start_date
            fee_date = fee.start_date.date() if hasattr(fee.start_date, 'date') else fee.start_date
            
            if schedule_date < fee_date:
                continue
        
        # 直接查询该课程安排的参与学员列表
        stmt = select(schedule_student.c.student_id, schedule_student.c.attendance_status).where(
            schedule_student.c.schedule_id == schedule.id
        )
        result = db.execute(stmt)
        scheduled_student_info = {row[0]: row[1] for row in result.fetchall()}
        if fee.student_id in scheduled_student_info and scheduled_student_info[fee.student_id] == 'present':
            # 检查是否已经创建了消耗日志
            existing_log = db.query(FeeLog).filter(
                FeeLog.student_id == fee.student_id,
                FeeLog.course_id == fee.course_id,
                FeeLog.schedule_id == schedule.id,
                FeeLog.log_type == "consume"
            ).first()
            
            if not existing_log:
                # 创建消耗日志
                start_time = datetime.strptime(schedule.start_time, "%H:%M")
                end_time = datetime.strptime(schedule.end_time, "%H:%M")
                hours = (end_time - start_time).total_seconds() / 3600
                
                fee.consumed_hours += hours
                fee.remaining_hours = fee.total_lesson_count * hpl_consume - fee.consumed_hours
                
                consume_log = FeeLog(
                    student_id=fee.student_id,
                    course_id=schedule.course_id,
                    schedule_id=schedule.id,
                    log_type="consume",
                    amount=-hours * fee.hourly_fee,
                    hours=hours,
                    remaining_amount=fee.total_actual_amount - fee.total_refund_amount,
                    remaining_hours=fee.remaining_hours,
                    description=f"完训消耗 {hours} 小时(根据规则和实际情况触发记录)"
                )
                db.add(consume_log)
                consumed_count += 1
    
    if consumed_count > 0:
        db.commit()
        db.refresh(fee)
        log_operation(db, "费用管理", "手动触发消耗", f"为学员 {student.name} - 科目 {course.name} 补录 {consumed_count} 条消耗记录", current_user.username)
        return {
            "message": f"成功补录 {consumed_count} 条消耗记录",
            "consumed_count": consumed_count,
            "remaining_hours": fee.remaining_hours
        }
    else:
        return {
            "message": "没有需要补录的消耗记录",
            "consumed_count": 0,
            "remaining_hours": fee.remaining_hours
        }

@router.put("/student-fees/{fee_id}", response_model=StudentFeeSchema)
def update_student_fee(
    fee_id: int,
    fee: StudentFeeUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")

    """更新学生课时费记录"""
    db_fee = db.query(StudentFee).filter(StudentFee.id == fee_id).first()
    if not db_fee:
        log_operation(db, "费用管理", "更新课时费记录失败", f"课时费记录ID {fee_id} 不存在", current_user.username, "WARNING")
        raise HTTPException(status_code=404, detail="课时费记录不存在")
    
    if fee.start_date is not None:
        db_fee.start_date = fee.start_date
    if fee.hourly_fee is not None:
        db_fee.hourly_fee = fee.hourly_fee
    if fee.alert_threshold is not None:
        db_fee.alert_threshold = fee.alert_threshold
    if fee.is_active is not None:
        db_fee.is_active = fee.is_active
    
    db.commit()
    log_operation(db, "费用管理", "修改", "成功更新课时费记录", current_user.username)
    db.refresh(db_fee)
    
    student = db.query(Student).filter(Student.id == db_fee.student_id).first()
    course = db.query(Course).filter(Course.id == db_fee.course_id).first()
    
    # 重新计算总实收金额：所有收费记录的实缴金额总和
    payment_logs = db.query(FeeLog).filter(
        FeeLog.student_id == db_fee.student_id,
        FeeLog.course_id == db_fee.course_id,
        FeeLog.log_type == "payment"
    ).all()
    calculated_total_actual_amount = sum(log.amount for log in payment_logs)
    
    # 计算总应收金额：所有收费记录的应收金额总和
    total_receivable_amount = sum(log.receivable_amount for log in payment_logs)
    
    # 计算总退费金额：所有退费记录的金额总和
    refund_logs = db.query(FeeLog).filter(
        FeeLog.student_id == db_fee.student_id,
        FeeLog.course_id == db_fee.course_id,
        FeeLog.log_type == "refund"
    ).all()
    total_refund_amount = sum(abs(log.amount) for log in refund_logs)
    
    # 计算已消耗课时
    consumed_logs = db.query(FeeLog).filter(
        FeeLog.student_id == db_fee.student_id,
        FeeLog.course_id == db_fee.course_id,
        FeeLog.log_type == "consume"
    ).all()
    consumed_hours = sum(log.hours for log in consumed_logs)
    
    # 获取消耗的课程安排列表
    consumed_schedules = []
    for log in consumed_logs:
        if log.schedule_id:
            schedule = db.query(Schedule).filter(Schedule.id == log.schedule_id).first()
            if schedule:
                teacher = db.query(Teacher).filter(Teacher.id == schedule.teacher_id).first()
                class_ = db.query(Class).filter(Class.id == schedule.class_id).first()
                room = db.query(Room).filter(Room.id == schedule.room_id).first()
                
                consumed_schedules.append({
                    "date": schedule.start_date.strftime("%Y-%m-%d") if schedule.start_date else "",
                    "time": f"{schedule.start_time}-{schedule.end_time}",
                    "teacher": teacher.name if teacher else "",
                    "class": class_.name if class_ else "",
                    "room": room.name if room else "",
                    "hours": log.hours
                })
    
    return StudentFeeSchema(
        id=db_fee.id,
        student_id=db_fee.student_id,
        course_id=db_fee.course_id,
        start_date=db_fee.start_date,
        hourly_fee=db_fee.hourly_fee,
        total_receivable_amount=total_receivable_amount,
        total_actual_amount=calculated_total_actual_amount,
        total_refund_amount=total_refund_amount,
        remaining_hours=db_fee.remaining_hours,
        consumed_hours=consumed_hours,
        consumed_schedules=consumed_schedules,
        alert_threshold=db_fee.alert_threshold,
        is_active=db_fee.is_active,
        student_name=student.name if student else "",
        course_name=course.name if course else "",
        created_at=db_fee.created_at,
        updated_at=db_fee.updated_at
    )

@router.delete("/student-fees/{fee_id}")
def delete_student_fee(
    fee_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")

    """删除学生课时费记录"""
    db_fee = db.query(StudentFee).filter(StudentFee.id == fee_id).first()
    if not db_fee:
        log_operation(db, "费用管理", "删除课时费记录失败", f"课时费记录ID {fee_id} 不存在", current_user.username, "WARNING")
        raise HTTPException(status_code=404, detail="课时费记录不存在")
    
    db.delete(db_fee)
    db.commit()
    log_operation(db, "费用管理", "删除", f"删除课时费记录: {db_fee.id}", current_user.username, "WARNING")
    return {"message": "删除成功"}


@router.post("/payments")
def add_payment(
    payment: PaymentRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")
    
    """添加收费记录（追缴）"""
    fee = db.query(StudentFee).filter(
        StudentFee.student_id == payment.student_id,
        StudentFee.course_id == payment.course_id
    ).first()
    
    if not fee:
        log_operation(db, "费用管理", "添加收费记录失败", f"学生ID {payment.student_id} - 科目ID {payment.course_id} 的课时费记录不存在", current_user.username, "WARNING")
        raise HTTPException(status_code=404, detail="课时费记录不存在")
    
    # 计算本次应收金额 = 课节数 * 小时数每课节 * (课时费/小时)
    hpl = get_hours_per_lesson(db)
    receivable_amount = payment.lesson_count * hpl * fee.hourly_fee
    actual_amount = receivable_amount - payment.discount_amount
    
    # 累加总应收金额
    fee.total_receivable_amount += receivable_amount
    # 累加总实收金额
    fee.total_actual_amount += actual_amount
    # 累加累计新增课节数
    fee.total_lesson_count += payment.lesson_count
    # 重新计算剩余课时 = 累计新增课节数 * 小时数每课节 - 已消耗课时数
    fee.remaining_hours = fee.total_lesson_count * hpl - fee.consumed_hours
    
    payment_method_desc = f"，{payment.payment_method}" if payment.payment_method else ""
    
    log = FeeLog(
        student_id=payment.student_id,
        course_id=payment.course_id,
        log_type="payment",
        amount=actual_amount,
        receivable_amount=receivable_amount,
        hours=0,
        remaining_amount=fee.total_actual_amount,
        remaining_hours=fee.remaining_hours,
        payment_date=payment.payment_date or date.today(),
        description=payment.description or f"追缴（{payment.lesson_count}节课，{payment.lesson_count * hpl}小时）{'，优惠' + str(payment.discount_amount) + '元' if payment.discount_amount > 0 else ''}{'，收费来自'+ payment_method_desc}"
    )
    db.add(log)
    db.commit()
    log_operation(db, "费用管理", "新增", f"成功追缴课费：{payment.lesson_count}节课，{actual_amount}元", current_user.username)
    db.refresh(fee)
    
    return {
        "message": "收费成功",
        "total_receivable_amount": fee.total_receivable_amount,
        "total_actual_amount": fee.total_actual_amount,
        "total_lesson_count": fee.total_lesson_count,
        "remaining_hours": fee.remaining_hours
    }


@router.post("/refunds")
def add_refund(
    refund: RefundRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")

    """添加退费记录"""
    fee = db.query(StudentFee).filter(
        StudentFee.student_id == refund.student_id,
        StudentFee.course_id == refund.course_id
    ).first()
    
    if not fee:
        log_operation(db, "费用管理", "添加退费记录失败", f"学生ID {refund.student_id} - 科目ID {refund.course_id} 的课时费记录不存在", current_user.username, "WARNING")
        raise HTTPException(status_code=404, detail="课时费记录不存在")
    
    # 计算退费金额对应的课时
    refunded_hours = refund.amount / fee.hourly_fee if fee.hourly_fee > 0 else 0
    
    # 累加总退费金额
    fee.total_refund_amount += refund.amount
    # 重新计算剩余课时 = 累计新增课节数 * 小时数每课节 - 已消耗课时数 - 退费课时
    fee.remaining_hours = fee.total_lesson_count * get_hours_per_lesson(db) - fee.consumed_hours - refunded_hours
    
    log = FeeLog(
        student_id=refund.student_id,
        course_id=refund.course_id,
        log_type="refund",
        amount=-refund.amount,
        hours=-refunded_hours,
        remaining_amount=fee.total_actual_amount - fee.total_refund_amount,
        remaining_hours=fee.remaining_hours,
        refund_date=refund.refund_date or date.today(),
        description=refund.refund_reason or refund.description or f"退费 {refund.amount} 元"
    )
    db.add(log)
    db.commit()
    log_operation(db, "费用管理", "新增", f"为学员{refund.student_id}退费：{refund.amount}元，原因：{refund.refund_reason}", current_user.username)
    db.refresh(fee)
    
    return {
        "message": "退费成功",
        "total_refund_amount": fee.total_refund_amount,
        "remaining_hours": fee.remaining_hours
    }


@router.post("/consume/{schedule_id}")
def consume_hours(
    schedule_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")

    """完训消耗课时（基于实际出勤学员）"""
    schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if not schedule:
        log_operation(db, "费用管理", "完训消耗课时失败", f"课程安排ID {schedule_id} 不存在", current_user.username, "WARNING")
        raise HTTPException(status_code=404, detail="课程安排不存在")
    
    query = select(schedule_student.c.student_id, schedule_student.c.attendance_status).where(
        schedule_student.c.schedule_id == schedule_id
    )
    result = db.execute(query)
    student_attendance = result.fetchall()
    
    if not student_attendance:
        log_operation(db, "费用管理", "完训消耗课时", f"课程安排ID {schedule_id} 没有学员记录", current_user.username, "WARNING")
        return []
    
    attended_students = [row[0] for row in student_attendance if row[1] == 'present']
    
    log_operation(db, "费用管理", "完训消耗课时", f"课程安排ID: {schedule_id}, 总学员数: {len(student_attendance)}, 出席学员数: {len(attended_students)}", current_user.username, "INFO")
    
    results = []
    hpl = get_hours_per_lesson(db)
    for student_id in attended_students:
        student = db.query(Student).filter(Student.id == student_id).first()
        if not student:
            continue
        fee = db.query(StudentFee).filter(
            StudentFee.student_id == student.id,
            StudentFee.course_id == schedule.course_id
        ).first()
        
        if not fee:
            log_operation(db, "费用管理", "完训消耗课时", f"学员 {student.name} (ID: {student.id}) 没有找到科目ID {schedule.course_id} 的课时费记录，跳过", current_user.username, "WARNING")
            continue
        
        if not fee.is_active:
            log_operation(db, "费用管理", "完训消耗课时", f"学员 {student.name} (ID: {student.id}) 的课时费记录未激活，跳过", current_user.username, "WARNING")
            continue
        
        start_time = datetime.strptime(schedule.start_time, "%H:%M")
        end_time = datetime.strptime(schedule.end_time, "%H:%M")
        hours = (end_time - start_time).total_seconds() / 3600
        
        fee.consumed_hours += hours
        fee.remaining_hours = fee.total_lesson_count * hpl - fee.consumed_hours
        
        log = FeeLog(
            student_id=student.id,
            course_id=schedule.course_id,
            schedule_id=schedule.id,
            log_type="consume",
            amount=-hours * fee.hourly_fee,
            hours=hours,
            remaining_amount=fee.total_actual_amount - fee.total_refund_amount,
            remaining_hours=fee.remaining_hours,
            description=f"完训消耗 {hours} 小时"
        )
        db.add(log)
        
        log_operation(db, "费用管理", "完训消耗课时", f"学员 {student.name} (ID: {student.id}) 消耗 {hours} 小时，剩余课时: {fee.remaining_hours}", current_user.username, "INFO")
        
        results.append({
            "student_id": student.id,
            "student_name": student.name,
            "remaining_hours": fee.remaining_hours
        })
    
    if not results:
        log_operation(db, "费用管理", "完训消耗课时", f"课程安排ID {schedule_id} 没有找到需要消耗课时的学员", current_user.username, "WARNING")
    
    db.commit()
    log_operation(db, "费用管理", "更新", f"因课程安排完训消耗课时，课程安排ID: {schedule_id}，共消耗 {len(results)} 个学员的课时", current_user.username)
    return results

@router.get("/fee-logs")
def get_fee_logs(
    skip: int = 0,
    limit: int = 100,
    student_id: Optional[int] = None,
    course_id: Optional[int] = None,
    log_type: Optional[str] = None,
    sort_field: Optional[str] = None,
    sort_order: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")
    
    """获取课时费日志列表（支持分页）"""
    query = db.query(FeeLog)
    
    if student_id is not None:
        query = query.filter(FeeLog.student_id == student_id)
    if course_id is not None:
        query = query.filter(FeeLog.course_id == course_id)
    if log_type is not None:
        query = query.filter(FeeLog.log_type == log_type)
    
    # 获取总数
    total = query.count()
    
    # 获取所有日志用于计算余额（按时间升序）
    all_logs_query = query.order_by(FeeLog.created_at.asc())
    all_logs = all_logs_query.all()
    
    # 计算每个学员每个科目的账户余额（按时间顺序）
    student_course_balance = {}
    log_balance_map = {}  # 存储每条日志对应的余额
    
    for log in all_logs:
        key = (log.student_id, log.course_id)
        if key not in student_course_balance:
            student_course_balance[key] = 0.0
        
        # 根据日志类型更新余额
        if log.log_type == 'payment':
            student_course_balance[key] += log.amount
        elif log.log_type == 'refund':
            student_course_balance[key] += log.amount
        elif log.log_type == 'consume':
            student_course_balance[key] += log.amount
        
        # 记录这条日志时的余额
        log_balance_map[log.id] = student_course_balance[key]
    
    # 应用排序和分页
    if sort_field and sort_order:
        if sort_field in ['created_at', 'student_name', 'course_name', 'log_type', 'amount', 'hours']:
            if sort_field == 'created_at':
                order_column = FeeLog.created_at
            elif sort_field == 'student_name':
                order_column = Student.name
            elif sort_field == 'course_name':
                order_column = Course.name
            elif sort_field == 'log_type':
                order_column = FeeLog.log_type
            elif sort_field == 'amount':
                order_column = FeeLog.amount
            elif sort_field == 'hours':
                order_column = FeeLog.hours
            
            if sort_order == 'asc':
                query = query.order_by(order_column.asc())
            else:
                query = query.order_by(order_column.desc())
    else:
        query = query.order_by(FeeLog.created_at.desc())
    
    # 分页查询
    logs = query.offset(skip).limit(limit).all()
    
    results = []
    for log in logs:
        # 获取学员和科目名称
        student = db.query(Student).filter(Student.id == log.student_id).first()
        course = db.query(Course).filter(Course.id == log.course_id).first()
        
        # 使用预先计算的余额
        student_balance = log_balance_map.get(log.id, 0.0)

        # 获取课程安排信息
        schedule_date = None
        schedule_time = None
        teacher_name = None
        class_name = None
        room_name = None
        content_feedback = None
        if log.schedule_id:
            schedule = db.query(Schedule).filter(Schedule.id == log.schedule_id).first()
            if schedule:
                schedule_date = schedule.start_date
                schedule_time = f"{schedule.start_time}-{schedule.end_time}"
                teacher = db.query(Teacher).filter(Teacher.id == schedule.teacher_id).first()
                teacher_name = teacher.name if teacher else None
                class_ = db.query(Class).filter(Class.id == schedule.class_id).first()
                class_name = class_.name if class_ else None
                room = db.query(Room).filter(Room.id == schedule.room_id).first()
                room_name = room.name if room else None
                content_feedback = schedule.content_feedback
 
        # 根据日志类型返回正确的日期
        log_date = log.created_at
        if log.log_type == 'payment' and log.payment_date:
            log_date = log.payment_date
        elif log.log_type == 'refund' and log.refund_date:
            log_date = log.refund_date
        elif log.log_type == 'consume' and schedule_date:
            log_date = schedule_date

        results.append(FeeLogSchema(
            id=log.id,
            student_id=log.student_id,
            course_id=log.course_id,
            schedule_id=log.schedule_id,
            log_type=log.log_type,
            amount=log.amount,
            receivable_amount=log.receivable_amount,
            hours=log.hours,
            remaining_amount=student_balance,
            remaining_hours=log.remaining_hours,
            description=log.description,
            created_at=log_date,
            student_name=student.name if student else "",
            course_name=course.name if course else "",
            schedule_date=schedule_date.strftime("%Y-%m-%d") if schedule_date else None,
            schedule_time=schedule_time,
            teacher_name=teacher_name,
            class_name=class_name,
            room_name=room_name,
            content_feedback=content_feedback
        ))
    
    return {
        "items": results,
        "total": total
    }

@router.get("/export")
def export_fees(
    student_id: Optional[int] = None,
    course_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")
 
    """导出课时费记录为Excel"""
    query = db.query(StudentFee)
    
    if student_id is not None:
        query = query.filter(StudentFee.student_id == student_id)
    if course_id is not None:
        query = query.filter(StudentFee.course_id == course_id)
    
    fees = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "课时费记录"
    
    headers = ["学生", "科目", "起算日期", "课时费/小时", "总应收金额", "总实收金额", "总退费金额", "累计课节数", "已消耗课时", "剩余课时", "预警阈值", "状态"]
    ws.append(headers)
    
    for fee in fees:
        student = db.query(Student).filter(Student.id == fee.student_id).first()
        course = db.query(Course).filter(Course.id == fee.course_id).first()
        
        payment_logs = db.query(FeeLog).filter(
            FeeLog.student_id == fee.student_id,
            FeeLog.course_id == fee.course_id,
            FeeLog.log_type == "payment"
        ).all()
        calculated_total_actual_amount = sum(log.amount for log in payment_logs)
        
        total_receivable_amount = sum(log.receivable_amount for log in payment_logs)
        
        refund_logs = db.query(FeeLog).filter(
            FeeLog.student_id == fee.student_id,
            FeeLog.course_id == fee.course_id,
            FeeLog.log_type == "refund"
        ).all()
        total_refund_amount = sum(abs(log.amount) for log in refund_logs)
        
        consumed_logs = db.query(FeeLog).filter(
            FeeLog.student_id == fee.student_id,
            FeeLog.course_id == fee.course_id,
            FeeLog.log_type == "consume"
        ).all()
        consumed_hours = sum(log.hours for log in consumed_logs)
        
        ws.append([
            student.name if student else "",
            course.name if course else "",
            fee.start_date.strftime("%Y-%m-%d") if fee.start_date else "",
            fee.hourly_fee,
            total_receivable_amount,
            calculated_total_actual_amount,
            total_refund_amount,
            fee.total_lesson_count,
            consumed_hours,
            fee.remaining_hours,
            fee.alert_threshold,
            "启用" if fee.is_active else "禁用"
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    log_operation(db, "费用管理", "导出", f"成功导出课时费记录，学生ID: {student_id}, 科目ID: {course_id}", current_user.username)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fees.xlsx"}
    )


@router.get("/alerts")
def get_fee_alerts(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")
    
    """获取收费提醒"""
    from models import Settings
    import json
    
    alerts = []
    
    # 获取设置
    settings = db.query(Settings).first()
    
    # 获取所有启用的课时费记录
    fees = db.query(StudentFee).filter(StudentFee.is_active == True).all()
    
    for fee in fees:
        # 检查剩余课时是否低于预警阈值
        if fee.remaining_hours <= fee.alert_threshold:
            student = db.query(Student).filter(Student.id == fee.student_id).first()
            course = db.query(Course).filter(Course.id == fee.course_id).first()
            
            alert_data = {
                "id": fee.id,
                "student_id": fee.student_id,
                "student_name": student.name if student else "",
                "course_id": fee.course_id,
                "course_name": course.name if course else "",
                "remaining_hours": fee.remaining_hours,
                "alert_threshold": fee.alert_threshold,
                "hourly_fee": fee.hourly_fee,
                "remaining_amount": fee.total_actual_amount - fee.total_refund_amount
            }
            
            alerts.append(alert_data)
            
            # 发送微信缴费预警到管理群
            try:
                # 构建微信消息内容
                content = f"""## 💰 课时费预警通知
> **学员：** {student.name if student else '未知'}
> **科目：** {course.name if course else '未知'}
> **剩余课时：** <font color="warning">{fee.remaining_hours:.2f} 小时</font>
> **剩余金额：** <font color="warning">¥{alert_data['remaining_amount']:.2f}</font>

请及时联系学员续费！"""
                wechat_notifier.send_message_by_type("fee_alert", content, is_markdown=True)
            except Exception as e:
                log_operation(db, "费用管理", "发送微信缴费预警失败", f"学员ID: {fee.student_id}, 科目ID: {fee.course_id}, 错误: {str(e)}", current_user.username, "ERROR")
            
            # 发送邮件提醒（如果启用）
            if settings and student and student.email:
                try:
                    email_notif_settings = json.loads(settings.email_notification_settings or "{}")
                    if email_notif_settings.get('homework_enabled', False):
                        email_config = json.loads(settings.email_config or '{}')
                        from utils.email_notifier import EmailNotifier
                        email_notifier = EmailNotifier()
                        email_notifier.load_config(settings.email_config or '{}', settings.site_name)
                        
                        # 构建邮件内容
                        subject = f"【{settings.site_name}】课时费预警通知"
                        html_content = f"""
                        <!DOCTYPE html>
                        <html>
                        <head>
                            <meta charset="UTF-8">
                            <style>
                                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                                .header {{ background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); color: white; padding: 20px; text-align: center; border-radius: 10px 10px 0 0; }}
                                .content {{ background: #f9f9f9; padding: 20px; border-radius: 0 0 10px 10px; }}
                                .warning {{ color: #f56c6c; font-weight: bold; }}
                                .info-item {{ margin: 10px 0; padding: 10px; background: white; border-left: 4px solid #409eff; }}
                            </style>
                        </head>
                        <body>
                            <div class="container">
                                <div class="header">
                                    <h2>💰 课时费预警通知</h2>
                                </div>
                                <div class="content">
                                    <div class="info-item">
                                        <strong>学员姓名：</strong>{student.name}
                                    </div>
                                    <div class="info-item">
                                        <strong>科目名称：</strong>{course.name if course else '未知'}
                                    </div>
                                    <div class="info-item">
                                        <strong>剩余课时：</strong><span class="warning">{fee.remaining_hours:.2f} 小时</span>
                                    </div>
                                    <div class="info-item">
                                        <strong>剩余金额：</strong><span class="warning">¥{alert_data['remaining_amount']:.2f}</span>
                                    </div>
                                    <div class="info-item">
                                        <strong>预警阈值：</strong>{fee.alert_threshold} 小时
                                    </div>
                                    <p style="margin-top: 20px; color: #909399;">
                                        温馨提示：您的课时即将用完，请及时联系机构续费，以免影响正常上课。
                                    </p>
                                </div>
                            </div>
                        </body>
                        </html>
                        """
                        
                        email_notifier.send_email([student.email], subject, html_content)
                        log_operation(db, "费用管理", "发送邮件缴费预警", f"已发送邮件给学员 {student.name} ({student.email})", current_user.username, "INFO")
                except Exception as e:
                    log_operation(db, "费用管理", "发送邮件缴费预警失败", f"学员ID: {fee.student_id}, 错误: {str(e)}", current_user.username, "ERROR")
    
    return alerts

@router.get("/export-payment-records")
def export_payment_records(
    student_id: Optional[int] = None,
    course_id: Optional[int] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")
    
    """导出学员收费记录"""
    from sqlalchemy.orm import joinedload
    
    query = db.query(StudentFee).options(
        joinedload(StudentFee.student).joinedload(Student.classes),
        joinedload(StudentFee.course).joinedload(Course.teachers)
    )
    
    if student_id is not None:
        query = query.filter(StudentFee.student_id == student_id)
    if course_id is not None:
        query = query.filter(StudentFee.course_id == course_id)
    if search:
        # 搜索学员姓名或科目名称
        query = query.join(Student).join(Course).filter(
            (Student.name.ilike(f'%{search}%')) | (Course.name.ilike(f'%{search}%'))
        )
    fees = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "课费管理"
    
    headers = ["学员", "科目", "起算日期", "课时费/小时", "累计应收金额", "累计实收金额", "累计退费金额", "当前累计收入", "累计已消耗课时", "当前剩余课时数", "预警阈值(小时)", "状态"]
    ws.append(headers)
    
    for fee in fees:
        student = fee.student
        course = fee.course
        
        # 计算总实收金额：所有收费记录的实缴金额总和
        payment_logs = db.query(FeeLog).filter(
            FeeLog.student_id == fee.student_id,
            FeeLog.course_id == fee.course_id,
            FeeLog.log_type == "payment"
        ).all()
        calculated_total_actual_amount = sum(log.amount for log in payment_logs)
        
        # 计算总应收金额：所有收费记录的应收金额总和
        total_receivable_amount = sum(log.receivable_amount for log in payment_logs)
        
        # 计算总退费金额：所有退费记录的金额总和
        refund_logs = db.query(FeeLog).filter(
            FeeLog.student_id == fee.student_id,
            FeeLog.course_id == fee.course_id,
            FeeLog.log_type == "refund"
        ).all()
        total_refund_amount = sum(abs(log.amount) for log in refund_logs)
        
        # 计算已消耗课时
        consumed_logs = db.query(FeeLog).filter(
            FeeLog.student_id == fee.student_id,
            FeeLog.course_id == fee.course_id,
            FeeLog.log_type == "consume"
        ).all()
        consumed_hours = sum(log.hours for log in consumed_logs)
        
        ws.append([
            student.name if student else "",
            course.name if course else "",
            fee.start_date.strftime("%Y-%m-%d") if fee.start_date else "",
            fee.hourly_fee,
            total_receivable_amount,
            calculated_total_actual_amount,
            total_refund_amount,
            calculated_total_actual_amount - total_refund_amount,
            consumed_hours,
            fee.remaining_hours,
            fee.alert_threshold,
            "启用" if fee.is_active else "禁用"
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    log_operation(db, "费用管理", "导出", f"成功导出学员收费记录，学生ID: {student_id}, 科目ID: {course_id}", current_user.username)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=payment-records.xlsx"}
    )

@router.get("/export-fee-logs")
def export_fee_logs(
    student_id: Optional[int] = None,
    course_id: Optional[int] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")
    
    """导出课时费记录"""
    from sqlalchemy.orm import joinedload
    
    query = db.query(FeeLog).options(
        joinedload(FeeLog.student),
        joinedload(FeeLog.course)
    )
    
    if student_id is not None:
        query = query.filter(FeeLog.student_id == student_id)
    if course_id is not None:
        query = query.filter(FeeLog.course_id == course_id)
    if search:
        # 搜索学员姓名或科目名称
        query = query.join(Student).join(Course).filter(
            (Student.name.contains(search)) | (Course.name.contains(search))
        )
    
    # 先按时间升序获取所有日志，用于正确计算余额
    logs_asc = query.order_by(FeeLog.created_at.asc()).all()
    
    # 获取学员课时费记录，用于获取最新的学员账户余额和剩余课时
    fee_query = db.query(StudentFee).options(
        joinedload(StudentFee.student),
        joinedload(StudentFee.course)
    )
    
    if student_id is not None:
        fee_query = fee_query.filter(StudentFee.student_id == student_id)
    if course_id is not None:
        fee_query = fee_query.filter(StudentFee.course_id == course_id)
    
    fees = fee_query.all()
    
    # 创建学员科目到课时费记录的映射
    student_fee_map = {}
    for fee in fees:
        key = (fee.student_id, fee.course_id)
        student_fee_map[key] = fee

    # 按时间顺序计算每个学员每个科目的账户余额
    student_course_balance = {}
    log_balance_map = {}  # 存储每条日志对应的余额
    
    for log in logs_asc:
        key = (log.student_id, log.course_id)
        if key not in student_course_balance:
            student_course_balance[key] = 0.0
        
        # 根据日志类型更新余额
        if log.log_type == 'payment':
            student_course_balance[key] += log.amount
        elif log.log_type == 'refund':
            student_course_balance[key] += log.amount
        elif log.log_type == 'consume':
            student_course_balance[key] += log.amount
        
        # 记录这条日志时的余额
        log_balance_map[log.id] = student_course_balance[key]
    
    # 按时间降序排列用于导出
    logs_desc = sorted(logs_asc, key=lambda x: x.created_at, reverse=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "课时费记录"
    
    headers = ["日期", "学员", "科目", "类型", "金额变化", "课时", "学员账户余额", "剩余课时", "描述"]
    ws.append(headers)
    
    for log in logs_desc:
        student = log.student
        course = log.course
        
        # 使用预先计算的余额
        student_balance = log_balance_map.get(log.id, 0.0)
        
        # 获取类型文本
        log_type_text = ""
        if log.log_type == 'payment':
            log_type_text = "缴费"
        elif log.log_type == 'refund':
            log_type_text = "退费"
        elif log.log_type == 'consume':
            log_type_text = "消耗"
        
        # 获取课程安排信息
        schedule_date = None
        schedule_time = None
        teacher_name = None
        class_name = None
        room_name = None
        content_feedback = None
        if log.schedule_id:
            schedule = db.query(Schedule).filter(Schedule.id == log.schedule_id).first()
            if schedule:
                schedule_date = schedule.start_date  # 保持为日期对象
                schedule_time = f"{schedule.start_time}-{schedule.end_time}"
                teacher = db.query(Teacher).filter(Teacher.id == schedule.teacher_id).first()
                teacher_name = teacher.name if teacher else None
                class_ = db.query(Class).filter(Class.id == schedule.class_id).first()
                class_name = class_.name if class_ else None
                room = db.query(Room).filter(Room.id == schedule.room_id).first()
                room_name = room.name if room else None
                content_feedback = schedule.content_feedback
        
        # 根据日志类型获取正确的日期
        log_date = log.created_at
        if log.log_type == 'payment' and log.payment_date:
            log_date = log.payment_date
        elif log.log_type == 'refund' and log.refund_date:
            log_date = log.refund_date
        elif log.log_type == 'consume' and schedule_date:
            log_date = schedule_date

        # 使用预先计算的账户余额
        ws.append([
            log_date.strftime("%Y-%m-%d") if log_date else "",
            student.name if student else "",
            course.name if course else "",
            log_type_text,
            f"{'+' if log.amount > 0 else ''}¥{log.amount:.2f}",
            f"{log.hours:.2f} 小时" if log.hours > 0 else '-',
            f"¥{student_balance:.2f}",
            f"{log.remaining_hours:.2f} 小时",
            log.description or ""
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    log_operation(db, "费用管理", "导出", f"成功导出课时费记录，学生ID: {student_id}, 科目ID: {course_id}", current_user.username)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fee-logs.xlsx"}
    )

def consume_hours_with_attendance(
    schedule_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")

    """完训消耗课时（基于实际出勤学员）"""
    schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if not schedule:
        log_operation(db, "费用管理", "完训消耗课时失败", f"课程安排ID {schedule_id} 不存在", current_user.username, "WARNING")
        raise HTTPException(status_code=404, detail="课程安排不存在")
    
    # 获取该课程安排的所有学员及其出勤状态
    query = select(schedule_student.c.student_id, schedule_student.c.attendance_status).where(
        schedule_student.c.schedule_id == schedule_id
    )
    result = db.execute(query)
    student_attendance = result.fetchall()
    
    if not student_attendance:
        log_operation(db, "费用管理", "完训消耗课时", f"课程安排ID {schedule_id} 没有学员记录", current_user.username, "WARNING")
        return []
    
    # 只处理出席的学员
    attended_students = [row[0] for row in student_attendance if row[1] == 'present']
    
    log_operation(db, "费用管理", "完训消耗课时", f"课程安排ID: {schedule_id}, 总学员数: {len(student_attendance)}, 出席学员数: {len(attended_students)}", current_user.username, "INFO")
    
    results = []
    hpl = get_hours_per_lesson(db)
    for student_id in attended_students:
        student = db.query(Student).filter(Student.id == student_id).first()
        if not student:
            continue
        
        fee = db.query(StudentFee).filter(
            StudentFee.student_id == student.id,
            StudentFee.course_id == schedule.course_id
        ).first()
        
        if not fee:
            log_operation(db, "费用管理", "完训消耗课时", f"学员 {student.name} (ID: {student.id}) 没有找到科目ID {schedule.course_id} 的课时费记录，跳过", current_user.username, "WARNING")
            continue
        
        if not fee.is_active:
            log_operation(db, "费用管理", "完训消耗课时", f"学员 {student.name} (ID: {student.id}) 的课时费记录未激活，跳过", current_user.username, "WARNING")
            continue
        
        start_time = datetime.strptime(schedule.start_time, "%H:%M")
        end_time = datetime.strptime(schedule.end_time, "%H:%M")
        hours = (end_time - start_time).total_seconds() / 3600
        
        fee.consumed_hours += hours
        fee.remaining_hours = fee.total_lesson_count * hpl - fee.consumed_hours
        
        log = FeeLog(
            student_id=student.id,
            course_id=schedule.course_id,
            schedule_id=schedule.id,
            log_type="consume",
            amount=-hours * fee.hourly_fee,
            hours=hours,
            remaining_amount=fee.total_actual_amount - fee.total_refund_amount,
            remaining_hours=fee.remaining_hours,
            description=f"完训消耗 {hours} 小时"
        )
        db.add(log)
        
        log_operation(db, "费用管理", "完训消耗课时", f"学员 {student.name} (ID: {student.id}) 消耗 {hours} 小时，剩余课时: {fee.remaining_hours}", current_user.username, "INFO")
        
        results.append({
            "student_id": student.id,
            "student_name": student.name,
            "remaining_hours": fee.remaining_hours
        })
    
    if not results:
        log_operation(db, "费用管理", "完训消耗课时", f"课程安排ID {schedule_id} 没有找到需要消耗课时的学员", current_user.username, "WARNING")
    
    db.commit()
    log_operation(db, "费用管理", "更新", f"因课程安排完训消耗课时，课程安排ID: {schedule_id}，共消耗 {len(results)} 个学员的课时", current_user.username)
    return results

def recalculate_consumed_hours(
    schedule_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")

    """重新计算课时消耗（当出勤状态改变时）"""
    schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if not schedule:
        log_operation(db, "费用管理", "重新计算课时失败", f"课程安排ID {schedule_id} 不存在", current_user.username, "WARNING")
        raise HTTPException(status_code=404, detail="课程安排不存在")
    
    if schedule.execution_status != "completed":
        log_operation(db, "费用管理", "重新计算课时", f"课程安排ID {schedule_id} 未完训，无需重新计算", current_user.username, "INFO")
        return []
    
    deleted_logs = db.query(FeeLog).filter(
        FeeLog.schedule_id == schedule_id,
        FeeLog.log_type == "consume"
    ).all()
    
    for log in deleted_logs:
        db.delete(log)
    
    # 重置所有相关学员的已消耗课时
    query = select(schedule_student.c.student_id, schedule_student.c.attendance_status).where(
        schedule_student.c.schedule_id == schedule_id
    )
    result = db.execute(query)
    all_students = result.fetchall()
    
    hpl = get_hours_per_lesson(db)
    # 先减去之前消耗的课时
    start_time = datetime.strptime(schedule.start_time, "%H:%M")
    end_time = datetime.strptime(schedule.end_time, "%H:%M")
    hours = (end_time - start_time).total_seconds() / 3600
    for log in deleted_logs:
        fee = db.query(StudentFee).filter(
            StudentFee.student_id == log.student_id,
            StudentFee.course_id == schedule.course_id
        ).first()
        
        if fee:
            fee.consumed_hours -= hours
            fee.remaining_hours = fee.total_lesson_count * hpl - fee.consumed_hours
    
    db.commit()
    log_operation(db, "费用管理", "重新计算课时", f"课程安排ID {schedule_id} 重新计算完成", current_user.username, "INFO")
    
    # 重新消耗出席学员的课时
    return consume_hours_with_attendance(schedule_id, db, current_user)

@router.post("/repair/fix-incorrect-consumption")
def fix_incorrect_consumption(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """一次性修复：扫描所有已完训课程，修复请假/缺席学员被错误消耗的课时"""
    if not check_fee_manager_permission(db, current_user):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员或费用管理导师权限")
    completed_schedules = db.query(Schedule).filter(
        Schedule.execution_status == 'completed'
    ).all()
    total_fixed = 0
    total_logs_deleted = 0
    total_hours_recovered = 0
    details = []
    for schedule in completed_schedules:
        stmt = select(
            schedule_student.c.student_id,
            schedule_student.c.attendance_status
        ).where(schedule_student.c.schedule_id == schedule.id)
        result = db.execute(stmt)
        student_attendance = result.fetchall()
        non_present = [row[0] for row in student_attendance if row[1] not in ('present', None)]
        if not non_present:
            continue
        incorrect_logs = db.query(FeeLog).filter(
            FeeLog.schedule_id == schedule.id,
            FeeLog.log_type == 'consume',
            FeeLog.student_id.in_(non_present)
        ).all()
        if not incorrect_logs:
            continue
        try:
            recalculate_consumed_hours(schedule.id, db, current_user)
            total_fixed += 1
            total_logs_deleted += len(incorrect_logs)
            for log in incorrect_logs:
                total_hours_recovered += log.hours
            details.append({
                "schedule_id": schedule.id,
                "schedule_date": str(schedule.start_date) if schedule.start_date else "",
                "course_name": schedule.course.name if schedule.course else "",
                "logs_deleted": len(incorrect_logs)
            })
        except Exception as e:
            db.rollback()
            details.append({
                "schedule_id": schedule.id,
                "error": str(e)
            })
    log_operation(db, "费用管理", "批量修复错误消耗",
                  f"修复 {total_fixed} 个课程安排，删除 {total_logs_deleted} 条错误日志，恢复 {total_hours_recovered} 小时",
                  current_user.username, "WARNING")
    return {
        "success": True,
        "fixed_schedules": total_fixed,
        "logs_deleted": total_logs_deleted,
        "hours_recovered": total_hours_recovered,
        "details": details
    }